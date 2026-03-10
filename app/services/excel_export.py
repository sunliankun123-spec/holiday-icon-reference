from __future__ import annotations

import io
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Iterable
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from PIL import Image as PILImage
import requests

PREVIEW_MAX_PX = 220
PREVIEW_COL_WIDTH = 34
PREVIEW_ROW_HEIGHT_PT = 170
RENDER_IMG_MAX_SIDE = 900


def _fit_size(width: int, height: int, max_px: int) -> tuple[int, int]:
    if width <= 0 or height <= 0:
        return max_px, max_px
    ratio = min(max_px / width, max_px / height, 1.0)
    return max(1, int(width * ratio)), max(1, int(height * ratio))


def _try_fetch_image(url: str, referer: str = "") -> tuple[bytes, str] | None:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    }
    if referer:
        headers["Referer"] = referer

    attempts = 3
    for idx in range(attempts):
        try:
            resp = requests.get(
                url,
                timeout=(10, 20),
                allow_redirects=True,
                headers=headers,
                verify=False,
            )
            if resp.status_code >= 400:
                continue
            content_type = resp.headers.get("content-type", "").lower()
            if not content_type.startswith("image/"):
                continue
            return resp.content, content_type
        except Exception:
            if idx == attempts - 1:
                return None
    return None


def _pinimg_size_fallback(url: str) -> str:
    if "pinimg.com" not in url:
        return url
    if "/originals/" in url:
        return url.replace("/originals/", "/736x/")
    return url


def _download_preview(url: str) -> str | None:
    if not url:
        return None
    try:
        fetched = _try_fetch_image(url, referer="https://www.pinterest.com/")
        if not fetched:
            fetched = _try_fetch_image(_pinimg_size_fallback(url), referer="https://www.pinterest.com/")
        if not fetched and "pinimg.com" in url:
            fetched = _try_fetch_image(url.replace("/originals/", "/564x/"), referer="https://www.pinterest.com/")
        # Some hosts block direct requests; image proxy fallback.
        if not fetched:
            proxy_url = f"https://wsrv.nl/?url={quote(url, safe='')}&n=-1"
            fetched = _try_fetch_image(proxy_url)
        if not fetched:
            proxy_url = f"https://images.weserv.nl/?url={quote(url, safe='')}&n=-1"
            fetched = _try_fetch_image(proxy_url)
        if not fetched:
            return None

        raw, content_type = fetched
        render_mode = bool(os.getenv("RENDER") or os.getenv("RENDER_EXTERNAL_URL"))
        content_main = content_type.split(";")[0].strip()

        # Always convert to JPEG so openpyxl can insert reliably.
        img = PILImage.open(io.BytesIO(raw)).convert("RGB")
        if render_mode:
            img.thumbnail((RENDER_IMG_MAX_SIDE, RENDER_IMG_MAX_SIDE))
        buf = io.BytesIO()
        jpeg_quality = 90 if render_mode else 95
        img.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
        raw = buf.getvalue()

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
        tmp.write(raw)
        tmp.flush()
        tmp.close()
        return tmp.name
    except Exception:
        return None


def write_excel(rows: Iterable[dict], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "icon_references"
    ws.append(["theme", "element", "image_index", "search_query", "source", "image_preview"])

    temp_files: list[str] = []
    embedded_count = 0
    failed_count = 0
    futures = {}
    render_mode = bool(os.getenv("RENDER") or os.getenv("RENDER_EXTERNAL_URL"))
    with ThreadPoolExecutor(max_workers=2 if render_mode else 8) as executor:
        for row_idx, row in enumerate(rows, start=2):
            ws.append(
                [
                    row["theme"],
                    row["element"],
                    row["image_index"],
                    row["search_query"],
                    row["source"],
                    "",
                ]
            )
            futures[executor.submit(_download_preview, row["image_url"])] = row_idx

        for future in as_completed(futures):
            row_idx = futures[future]
            temp_path = future.result()
            if not temp_path:
                failed_count += 1
                continue
            temp_files.append(temp_path)
            img = XLImage(temp_path)
            img.width, img.height = _fit_size(img.width, img.height, PREVIEW_MAX_PX)
            ws.add_image(img, f"F{row_idx}")
            ws.row_dimensions[row_idx].height = PREVIEW_ROW_HEIGHT_PT
            embedded_count += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = PREVIEW_COL_WIDTH
    ws.freeze_panes = "A2"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = center

    ws2 = wb.create_sheet(title="element_list")
    ws2.append(["theme", "element"])
    seen = set()
    for row in rows:
        key = (row["theme"], row["element"])
        if key in seen:
            continue
        seen.add(key)
        ws2.append([row["theme"], row["element"]])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = center

    ws3 = wb.create_sheet(title="export_report")
    ws3.append(["metric", "value"])
    ws3.append(["total_rows", len(rows)])
    ws3.append(["embedded_images", embedded_count])
    ws3.append(["failed_images", failed_count])

    wb.save(output_path)
    for temp_file in temp_files:
        try:
            Path(temp_file).unlink(missing_ok=True)
        except Exception:
            pass

    return output_path
